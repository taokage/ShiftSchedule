import numpy as np
from openpyxl import load_workbook
from tkinter import Tk, filedialog


NUM_DAYS = 31
NUM_STAFF = 30
NUM_SHIFTS = 2


def select_excel_file():
    root = Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="Excelファイルを選択してください",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )

    root.destroy()

    if not file_path:
        raise FileNotFoundError("Excelファイルが選択されませんでした。")

    return file_path


def load_shift_schedule_excel():
    file_path = select_excel_file()
    wb = load_workbook(file_path, data_only=True, read_only=True)

    print("読み込んだファイル:", file_path)
    print("シート一覧:")
    for s in wb.sheetnames:
        print(repr(s))

    unavailable_3d = np.zeros((NUM_STAFF, NUM_DAYS, NUM_SHIFTS), dtype=bool)
    avoid_3d = np.zeros((NUM_STAFF, NUM_DAYS, NUM_SHIFTS), dtype=bool)
    mandatory_3d = np.zeros((NUM_STAFF, NUM_DAYS, NUM_SHIFTS), dtype=bool)

    ew1_mandatory_3d = np.zeros((NUM_STAFF, NUM_DAYS, NUM_SHIFTS), dtype=bool)
    iw1_mandatory_3d = np.zeros((NUM_STAFF, NUM_DAYS, NUM_SHIFTS), dtype=bool)

    optimal_staff_2d = np.zeros((NUM_STAFF, NUM_SHIFTS), dtype=int)
    min_work_count_2d = np.zeros((NUM_DAYS, NUM_SHIFTS), dtype=int)
    min_leader_count_2d = np.zeros((NUM_DAYS, NUM_SHIFTS), dtype=int)

    ew1_count_2d = np.zeros((NUM_DAYS, NUM_SHIFTS), dtype=int)
    iw1_count_2d = np.zeros((NUM_DAYS, NUM_SHIFTS), dtype=int)

    day_week_2d = np.empty((NUM_DAYS, 2), dtype=object)

    for shift in range(NUM_SHIFTS):
        unavailable_sheet = wb[f"unavailable_3d_shift{shift}"]
        avoid_sheet = wb[f"avoid_3d_shift{shift}"]
        mandatory_sheet = wb[f"mandatory_3d_shift{shift}"]
        ew1_mandatory_sheet = wb[f"ew1_mandatory_3d_shift{shift}"]
        iw1_mandatory_sheet = wb[f"iw1_mandatory_3d_shift{shift}"]

        for day in range(NUM_DAYS):
            for staff in range(NUM_STAFF):
                unavailable_3d[staff, day, shift] = bool(
                    unavailable_sheet.cell(row=day + 1, column=staff + 1).value
                )
                avoid_3d[staff, day, shift] = bool(
                    avoid_sheet.cell(row=day + 1, column=staff + 1).value
                )
                mandatory_3d[staff, day, shift] = bool(
                    mandatory_sheet.cell(row=day + 1, column=staff + 1).value
                )
                ew1_mandatory_3d[staff, day, shift] = bool(
                    ew1_mandatory_sheet.cell(row=day + 1, column=staff + 1).value
                )
                iw1_mandatory_3d[staff, day, shift] = bool(
                    iw1_mandatory_sheet.cell(row=day + 1, column=staff + 1).value
                )

    for shift in range(NUM_SHIFTS):
        ws = wb[f"optimal_staff_2d_shift{shift}"]
        for staff in range(NUM_STAFF):
            v = ws.cell(row=1, column=staff + 1).value
            optimal_staff_2d[staff, shift] = 0 if v in [None, ""] else int(v)

    for shift in range(NUM_SHIFTS):
        ws = wb[f"min_work_count_2d_shift{shift}"]
        for day in range(NUM_DAYS):
            v = ws.cell(row=day + 1, column=1).value
            min_work_count_2d[day, shift] = 0 if v in [None, ""] else int(v)

    for shift in range(NUM_SHIFTS):
        ws = wb[f"min_leader_count_2d_shift{shift}"]
        for day in range(NUM_DAYS):
            v = ws.cell(row=day + 1, column=1).value
            min_leader_count_2d[day, shift] = 0 if v in [None, ""] else int(v)

    for shift in range(NUM_SHIFTS):
        ws = wb[f"ew1_count_2d_shift{shift}"]
        for day in range(NUM_DAYS):
            v = ws.cell(row=day + 1, column=1).value
            ew1_count_2d[day, shift] = 0 if v in [None, ""] else int(v)

    for shift in range(NUM_SHIFTS):
        ws = wb[f"iw1_count_2d_shift{shift}"]
        for day in range(NUM_DAYS):
            v = ws.cell(row=day + 1, column=1).value
            iw1_count_2d[day, shift] = 0 if v in [None, ""] else int(v)

    ws_day_week = wb["day_week_2d"]
    for day in range(NUM_DAYS):
        day_week_2d[day, 0] = ws_day_week.cell(row=day + 1, column=1).value
        day_week_2d[day, 1] = ws_day_week.cell(row=day + 1, column=2).value

    ws_staffno = wb["staffno_1d"]
    staffno_1d = np.zeros(NUM_STAFF, dtype=int)
    for staff in range(NUM_STAFF):
        v = ws_staffno.cell(row=1, column=staff + 1).value
        staffno_1d[staff] = 0 if v in [None, ""] else int(v)

    ws_staffname = wb["staffname_1d"]
    staffname_1d = np.empty(NUM_STAFF, dtype=object)
    for staff in range(NUM_STAFF):
        v = ws_staffname.cell(row=1, column=staff + 1).value
        staffname_1d[staff] = "" if v in [None, ""] else str(v).strip()

    # leader_level_1d
    # 0: リーダー不可
    # 1: リーダーなりたて
    # 2: 十分な力量のリーダー
    ws_leader_level = wb["leader_level_1d"]
    leader_level_1d = np.zeros(NUM_STAFF, dtype=int)
    for staff in range(NUM_STAFF):
        v = ws_leader_level.cell(row=1, column=staff + 1).value
        leader_level_1d[staff] = 0 if v in [None, ""] else int(v)

    ws_is_ew1_candidate = wb["is_ew1_candidate_1d"]
    is_ew1_candidate_1d = np.zeros(NUM_STAFF, dtype=bool)
    for staff in range(NUM_STAFF):
        v = ws_is_ew1_candidate.cell(row=1, column=staff + 1).value
        is_ew1_candidate_1d[staff] = bool(v)

    ws_iw1_priority = wb["iw1_priority_1d"]
    iw1_priority_1d = np.zeros(NUM_STAFF, dtype=int)
    for staff in range(NUM_STAFF):
        v = ws_iw1_priority.cell(row=1, column=staff + 1).value
        iw1_priority_1d[staff] = 0 if v in [None, ""] else int(v)

    night_pair_ng_2d = np.zeros((NUM_STAFF, NUM_STAFF), dtype=bool)

    staffname_to_index = {
        str(name).strip(): staff
        for staff, name in enumerate(staffname_1d)
        if str(name).strip() != ""
    }

    if "night_pair_ng" in wb.sheetnames:
        ws_ng = wb["night_pair_ng"]

        for row in range(1, ws_ng.max_row + 1):
            name1 = ws_ng.cell(row=row, column=2).value
            name2 = ws_ng.cell(row=row, column=3).value

            if name1 in [None, ""] or name2 in [None, ""]:
                continue

            name1 = str(name1).strip()
            name2 = str(name2).strip()

            if name1 not in staffname_to_index:
                print(f"night_pair_ng: 未登録スタッフ名です: {name1}")
                continue

            if name2 not in staffname_to_index:
                print(f"night_pair_ng: 未登録スタッフ名です: {name2}")
                continue

            staff1 = staffname_to_index[name1]
            staff2 = staffname_to_index[name2]

            if staff1 == staff2:
                continue

            night_pair_ng_2d[staff1, staff2] = True
            night_pair_ng_2d[staff2, staff1] = True
    else:
        print("night_pair_ng シートがありません。night_pair_ng_2d は全てFalseです。")

    debug_print_loaded_data(
        unavailable_3d,
        avoid_3d,
        mandatory_3d,
        ew1_mandatory_3d,
        iw1_mandatory_3d,
        optimal_staff_2d,
        min_work_count_2d,
        min_leader_count_2d,
        ew1_count_2d,
        iw1_count_2d,
        staffno_1d,
        staffname_1d,
        leader_level_1d,
        is_ew1_candidate_1d,
        iw1_priority_1d,
        day_week_2d,
        night_pair_ng_2d,
    )

    return (
        unavailable_3d,
        avoid_3d,
        mandatory_3d,
        ew1_mandatory_3d,
        iw1_mandatory_3d,
        optimal_staff_2d,
        min_work_count_2d,
        min_leader_count_2d,
        ew1_count_2d,
        iw1_count_2d,
        staffno_1d,
        staffname_1d,
        leader_level_1d,
        is_ew1_candidate_1d,
        iw1_priority_1d,
        day_week_2d,
        night_pair_ng_2d,
        file_path,
    )


def debug_print_loaded_data(
    unavailable_3d,
    avoid_3d,
    mandatory_3d,
    ew1_mandatory_3d,
    iw1_mandatory_3d,
    optimal_staff_2d,
    min_work_count_2d,
    min_leader_count_2d,
    ew1_count_2d,
    iw1_count_2d,
    staffno_1d,
    staffname_1d,
    leader_level_1d,
    is_ew1_candidate_1d,
    iw1_priority_1d,
    day_week_2d,
    night_pair_ng_2d,
):
    print("===== LOAD CHECK =====")

    print(f"unavailable_3d shape: {unavailable_3d.shape}")
    print(f"avoid_3d shape: {avoid_3d.shape}")
    print(f"mandatory_3d shape: {mandatory_3d.shape}")
    print(f"ew1_mandatory_3d shape: {ew1_mandatory_3d.shape}")
    print(f"iw1_mandatory_3d shape: {iw1_mandatory_3d.shape}")

    print(f"optimal_staff_2d shape: {optimal_staff_2d.shape}")
    print(f"min_work_count_2d shape: {min_work_count_2d.shape}")
    print(f"min_leader_count_2d shape: {min_leader_count_2d.shape}")
    print(f"ew1_count_2d shape: {ew1_count_2d.shape}")
    print(f"iw1_count_2d shape: {iw1_count_2d.shape}")

    print(f"staffno_1d shape: {staffno_1d.shape}")
    print(f"staffname_1d shape: {staffname_1d.shape}")
    print(f"leader_level_1d shape: {leader_level_1d.shape}")
    print(f"is_ew1_candidate_1d shape: {is_ew1_candidate_1d.shape}")
    print(f"iw1_priority_1d shape: {iw1_priority_1d.shape}")
    print(f"day_week_2d shape: {day_week_2d.shape}")
    print(f"night_pair_ng_2d shape: {night_pair_ng_2d.shape}")

    print(f"unavailable count: {np.sum(unavailable_3d)}")
    print(f"avoid count: {np.sum(avoid_3d)}")
    print(f"mandatory count: {np.sum(mandatory_3d)}")
    print(f"ew1 mandatory count: {np.sum(ew1_mandatory_3d)}")
    print(f"iw1 mandatory count: {np.sum(iw1_mandatory_3d)}")

    print(f"leader level 1 count: {np.sum(leader_level_1d == 1)}")
    print(f"leader level 2 count: {np.sum(leader_level_1d == 2)}")
    print(f"leader level >=1 count: {np.sum(leader_level_1d >= 1)}")

    print(f"ew1 candidate count: {np.sum(is_ew1_candidate_1d)}")
    print(f"night pair ng count: {np.sum(night_pair_ng_2d) // 2}")

    print("staffname_1d:")
    for i, name in enumerate(staffname_1d):
        print(f"{i}: {name}")

    print("leader_level_1d:")
    for i, level in enumerate(leader_level_1d):
        print(f"{i}: {level}")

    print("iw1_priority_1d:")
    for i, priority in enumerate(iw1_priority_1d):
        print(f"{i}: {priority}")

    print("===== LOAD COMPLETE =====")