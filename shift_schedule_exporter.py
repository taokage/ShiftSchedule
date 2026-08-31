from openpyxl import Workbook


NUM_STAFF = 30
NUM_DAYS = 31
NUM_SHIFTS = 2

DAY_SHIFT = 0
NIGHT_SHIFT = 1


def export_final_work_3d(
    final_work_3d,
    final_ew1_work_3d,
    final_iw1_work_3d,
    unavailable_3d,
    staffno_1d=None,
    staffname_1d=None,
    day_week_2d=None,
    file_path="shift_schedule_result.xlsx",
):
    wb = Workbook()

    ws = wb.active
    ws.title = "final_work_sheet"

    for day in range(NUM_DAYS):
        row = 6 + day

        if day_week_2d is not None:
            ws.cell(row=row, column=1).value = day_week_2d[day, 0]
            ws.cell(row=row, column=2).value = day_week_2d[day, 1]
        else:
            ws.cell(row=row, column=1).value = day + 1
            ws.cell(row=row, column=2).value = ""

    for staff in range(NUM_STAFF):
        start_col = 3 + staff * 2

        if staffno_1d is not None:
            ws.cell(row=4, column=start_col).value = int(staffno_1d[staff])
        else:
            ws.cell(row=4, column=start_col).value = staff

        if staffname_1d is not None:
            ws.cell(row=5, column=start_col).value = str(staffname_1d[staff])
        else:
            ws.cell(row=5, column=start_col).value = ""

        for day in range(NUM_DAYS):
            row = 6 + day

            if unavailable_3d[staff, day, DAY_SHIFT]:
                ws.cell(row=row, column=start_col).value = "×"

            if unavailable_3d[staff, day, NIGHT_SHIFT]:
                ws.cell(row=row, column=start_col + 1).value = "×"

            if final_work_3d[staff, day, DAY_SHIFT]:
                ws.cell(row=row, column=start_col).value = 1

            if final_work_3d[staff, day, NIGHT_SHIFT]:
                ws.cell(row=row, column=start_col + 1).value = 2

            if final_ew1_work_3d[staff, day, DAY_SHIFT]:
                ws.cell(row=row, column=start_col).value = "外1日"

            if final_ew1_work_3d[staff, day, NIGHT_SHIFT]:
                ws.cell(row=row, column=start_col + 1).value = "外1夜"

            if final_iw1_work_3d[staff, day, DAY_SHIFT]:
                ws.cell(row=row, column=start_col).value = "内1日"

            if final_iw1_work_3d[staff, day, NIGHT_SHIFT]:
                ws.cell(row=row, column=start_col + 1).value = "内1夜"

    wb.save(file_path)

    return file_path