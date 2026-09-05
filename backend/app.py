from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile
from datetime import date, datetime, timezone
import os

import numpy as np
from fastapi import Body, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from ortools.sat.python import cp_model
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import JSON, DateTime, String, create_engine, select, text
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker

from shift_schedule_exporter import export_final_work_3d
from shift_schedule_loader import load_shift_schedule_excel
from shift_schedule_solver import solve_shift_schedule

N_STAFF, N_DAYS, N_SHIFTS = 30, 31, 2

REQUEST_CHOICES = [
    ("unavailable", "×"),
    ("avoid", "△"),
    ("available", "○"),
    ("mandatory", "勤務"),
    ("want", "勤務希望"),
    ("research", "研究日希望"),
    ("regular_outside", "通常外勤"),
    ("ew1", "西大寺勤務"),
    ("ew2", "薬師寺勤務"),
    ("ew3", "吉備勤務"),
    ("iw1", "クリクラ"),
]
REQUEST_TO_LABEL = dict(REQUEST_CHOICES)
LABEL_TO_REQUEST = {label: value for value, label in REQUEST_CHOICES}

class StaffInput(BaseModel):
    no: int = 0
    name: str = ""
    emergency_count: int = Field(0, ge=0, le=1)
    leader_level: int = Field(0, ge=0, le=2)
    # EW1/IW1 は現行solverに接続。EW2/IW2 は管理情報として保持し、
    # 将来solverを拡張するときにそのまま利用できるようAPIでも受け取る。
    ew1_candidate: bool = False
    ew1_available: int = Field(0, ge=0, le=1)
    ew2_available: int = Field(0, ge=0, le=1)
    ew3_available: int = Field(0, ge=0, le=1)
    # IW1は0=対象外、1〜3=優先度を兼ねた整数値として扱う。
    iw1_available: int = Field(0, ge=0, le=3)
    iw2_available: int = Field(0, ge=0, le=1)
    target_day: int = Field(0, ge=0, le=20)
    target_night: int = Field(0, ge=0, le=10)

class ScheduleInput(BaseModel):
    dates: list[str]
    staff: list[StaffInput]
    requests: list[list[list[str]]]
    coverage: list[list[dict[str, int]]]
    night_pair_ng: list[list[int]] = []
    @field_validator("dates")
    @classmethod
    def dates_are_31(cls, value):
        if len(value) != N_DAYS: raise ValueError("日付は31日分必要です。")
        return value
    @field_validator("staff")
    @classmethod
    def staff_are_30_or_less(cls, value):
        if not 1 <= len(value) <= N_STAFF: raise ValueError("スタッフは1〜30名で入力してください。")
        return value

class RequestExcelInput(BaseModel):
    dates: list[str]
    staff: list[StaffInput]
    requests: list[list[list[str]]]
    remarks: list[list[list[str]]]
    coverage: list[list[dict[str, int]]] = []
    holiday_labels: list[str] = []

    @field_validator("dates")
    @classmethod
    def request_dates_are_31(cls, value):
        if len(value) != N_DAYS:
            raise ValueError("日付は31日分必要です。")
        return value

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql+psycopg://shift_user:shift_password@db:5432/shift_schedule",
)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)


class Base(DeclarativeBase):
    pass


class AppState(Base):
    __tablename__ = "app_state"

    key: Mapped[str] = mapped_column(String(100), primary_key=True)
    value: Mapped[dict | list] = mapped_column(JSON, nullable=False)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        default=lambda: datetime.now(timezone.utc),
        onupdate=lambda: datetime.now(timezone.utc),
        nullable=False,
    )


class StatePayload(BaseModel):
    value: dict | list


app = FastAPI(title="Shift Schedule API", version="3.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def create_database_tables():
    Base.metadata.create_all(bind=engine)


@app.get("/")
def root():
    return {"message": "Shift Schedule Backend", "status": "running", "database": "postgresql"}


@app.get("/health")
def health():
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return {"status": "ok", "database": "ok"}
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"database unavailable: {exc}") from exc


@app.get("/state/{key}")
def get_state(key: str):
    with SessionLocal() as session:
        row = session.get(AppState, key)
        if row is None:
            raise HTTPException(status_code=404, detail="state not found")
        return {"key": row.key, "value": row.value, "updated_at": row.updated_at}


@app.put("/state/{key}")
def put_state(key: str, payload: StatePayload):
    with SessionLocal() as session:
        row = session.get(AppState, key)
        if row is None:
            row = AppState(key=key, value=payload.value)
            session.add(row)
        else:
            row.value = payload.value
            row.updated_at = datetime.now(timezone.utc)
        session.commit()
        return {"saved": True, "key": key}

def arrays_from_payload(payload):
    unavailable = np.ones((N_STAFF, N_DAYS, N_SHIFTS), dtype=bool)
    avoid = np.zeros_like(unavailable); mandatory = np.zeros_like(unavailable)
    ew1_mandatory = np.zeros_like(unavailable); iw1_mandatory = np.zeros_like(unavailable)
    optimal = np.zeros((N_STAFF, N_SHIFTS), dtype=int)
    leader = np.zeros(N_STAFF, dtype=int); ew1_candidate = np.zeros(N_STAFF, dtype=bool)
    iw1_priority = np.zeros(N_STAFF, dtype=int); staffno = np.arange(1, N_STAFF + 1, dtype=int)
    staffname = np.full(N_STAFF, "", dtype=object)
    for s, member in enumerate(payload.staff):
        staffno[s] = member.no or s + 1; staffname[s] = member.name.strip()
        leader[s] = member.leader_level
        ew1_candidate[s] = bool(member.ew1_available or member.ew1_candidate)
        # IW1の0〜3をそのままsolver内部の優先度配列へ渡す。
        iw1_priority[s] = int(member.iw1_available)
        optimal[s] = [member.target_day, member.target_night]
        if member.name.strip():
            unavailable[s, :, :] = False
            for d in range(min(N_DAYS, len(payload.requests[s]))):
                for sh in range(N_SHIFTS):
                    state = payload.requests[s][d][sh]
                    unavailable[s,d,sh] = state == "unavailable"
                    avoid[s,d,sh] = state == "avoid"; mandatory[s,d,sh] = state == "mandatory"
    min_work = np.zeros((N_DAYS,N_SHIFTS),dtype=int); min_leader=np.zeros_like(min_work)
    ew1_count=np.zeros_like(min_work); iw1_count=np.zeros_like(min_work)
    for d, shifts in enumerate(payload.coverage[:N_DAYS]):
        for sh, item in enumerate(shifts[:N_SHIFTS]):
            min_work[d,sh]=max(0,int(item.get("minimum",0))); min_leader[d,sh]=max(0,int(item.get("leaders",0)))
            ew1_count[d,sh]=max(0,int(item.get("ew1",0))); iw1_count[d,sh]=max(0,int(item.get("iw1",0)))
    day_week=np.empty((N_DAYS,2),dtype=object); weekdays="月火水木金土日"
    for d, raw in enumerate(payload.dates):
        parsed=date.fromisoformat(raw); day_week[d]=[raw,weekdays[parsed.weekday()]]
    pair_ng=np.zeros((N_STAFF,N_STAFF),dtype=bool)
    for pair in payload.night_pair_ng:
        if len(pair)==2 and all(0<=x<N_STAFF for x in pair): pair_ng[pair[0],pair[1]]=pair_ng[pair[1],pair[0]]=True
    return (unavailable,avoid,mandatory,ew1_mandatory,iw1_mandatory,optimal,min_work,min_leader,ew1_count,iw1_count,staffno,staffname,leader,ew1_candidate,iw1_priority,day_week,pair_ng)

def solve_payload(payload):
    a=arrays_from_payload(payload)
    unavailable,avoid,mandatory,ew1m,iw1m,optimal,min_work,min_leader,ew1c,iw1c,staffno,staffname,leader,is_ew1,iw1p,day_week,pair_ng=a
    work,ew1,iw1,status=solve_shift_schedule(unavailable,avoid,mandatory,ew1m,iw1m,optimal,min_work,min_leader,ew1c,iw1c,staffno,leader,is_ew1,iw1p,pair_ng)
    if status not in (cp_model.OPTIMAL,cp_model.FEASIBLE): raise HTTPException(422,"条件を満たす勤務表が見つかりません。必要人数や勤務不可を確認してください。")
    return a,work,ew1,iw1


def _request_excel_workbook(payload: RequestExcelInput):
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務希望入力"
    option_ws = wb.create_sheet("選択肢")
    for row, (_, label) in enumerate(REQUEST_CHOICES, start=1):
        option_ws.cell(row=row, column=1, value=label)
    option_ws.sheet_state = "hidden"

    thin = Side(style="thin", color="D6DCE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 勤務者と勤務者の境界線
    staff_separator = Side(style="medium", color="000000")

    header_fill = PatternFill("solid", fgColor="E9EEF5")
    sub_fill = PatternFill("solid", fgColor="F5F7FA")
    holiday_fill = PatternFill("solid", fgColor="FFF1F1")
    remark_fill = PatternFill("solid", fgColor="FAFAFA")
    locked_fill = PatternFill("solid", fgColor="F1F3F5")
    pink_fill = PatternFill("solid", fgColor="FCE4EC")

    ws["A1"] = "勤務者ID"
    ws["A2"] = "勤務者名"
    ws["A3"] = "日勤・夜勤"
    for row in range(1, 4):
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    for i in range(N_STAFF):
        col = 5 + i * 2
        name = payload.staff[i].name if i < len(payload.staff) else ""
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        ws.cell(1, col, i)
        ws.cell(2, col, name)
        ws.cell(3, col, "日勤")
        ws.cell(3, col + 1, "夜勤")
        for row in (1, 2):
            cell = ws.cell(row, col)
            cell.fill = header_fill
            cell.border = border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in (col, col + 1):
            ws.cell(3, c).fill = sub_fill
            ws.cell(3, c).border = border
            ws.cell(3, c).font = Font(bold=True)
            ws.cell(3, c).alignment = Alignment(horizontal="center", vertical="center")

        # 勤務者名が入力されている場合は、氏名セルを薄いピンクにする。
        name_coord = ws.cell(2, col).coordinate
        ws.conditional_formatting.add(
            f"{name_coord}:{ws.cell(2, col + 1).coordinate}",
            FormulaRule(formula=[f'LEN({name_coord})>0'], fill=pink_fill),
        )

    dropdown = DataValidation(
        type="list",
        formula1="'選択肢'!$A$1:$A$11",
        allow_blank=False,
    )
    dropdown.error = "一覧から勤務希望を選択してください。"
    dropdown.errorTitle = "勤務希望の入力エラー"
    dropdown.prompt = "勤務希望をドロップダウンから選択してください。"
    dropdown.promptTitle = "勤務希望"
    ws.add_data_validation(dropdown)

    weekdays = "月火水木金土日"
    for d, raw_date in enumerate(payload.dates):
        request_row = 4 + d * 2
        remark_row = request_row + 1
        parsed = date.fromisoformat(raw_date)
        holiday_label = (
            payload.holiday_labels[d]
            if d < len(payload.holiday_labels) and payload.holiday_labels[d] in {"平日", "休日"}
            else ("休日" if parsed.weekday() >= 5 else "平日")
        )
        ws.cell(request_row, 1, d)
        ws.cell(request_row, 2, parsed)
        ws.cell(request_row, 3, weekdays[parsed.weekday()])
        ws.cell(request_row, 4, holiday_label)
        ws.cell(remark_row, 4, "備考")
        ws.cell(request_row, 2).number_format = "yyyy/mm/dd"

        base_fill = holiday_fill if holiday_label == "休日" else PatternFill(fill_type=None)
        for col in range(1, 5):
            ws.cell(request_row, col).border = border
            ws.cell(request_row, col).fill = base_fill
            ws.cell(request_row, col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(remark_row, col).border = border
            ws.cell(remark_row, col).fill = remark_fill
            ws.cell(remark_row, col).alignment = Alignment(horizontal="center", vertical="center")

        for i in range(N_STAFF):
            for sh in range(N_SHIFTS):
                col = 5 + i * 2 + sh
                state = "available"
                if i < len(payload.requests) and d < len(payload.requests[i]) and sh < len(payload.requests[i][d]):
                    state = payload.requests[i][d][sh]
                label = REQUEST_TO_LABEL.get(state, "○")
                ws.cell(request_row, col, label)
                remark = ""
                if i < len(payload.remarks) and d < len(payload.remarks[i]) and sh < len(payload.remarks[i][d]):
                    remark = payload.remarks[i][d][sh] or ""
                ws.cell(remark_row, col, remark)
                ws.cell(request_row, col).border = border
                ws.cell(request_row, col).fill = base_fill
                ws.cell(request_row, col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(remark_row, col).border = border
                ws.cell(remark_row, col).fill = remark_fill
                ws.cell(remark_row, col).alignment = Alignment(horizontal="left", vertical="center")
                dropdown.add(ws.cell(request_row, col))

    # --------------------------------------------------
    # BM列以降：日別の管理者設定
    # BM:BN 最小勤務数 / BO:BP 最小リーダー数
    # BQ:BR EW1 / BS:BT EW2 / BU:BV EW3 / BW:BX IW1 / BY:BZ IW2
    # --------------------------------------------------
    coverage_sections = [
        (65, "最小勤務数", "minimum"),
        (67, "最小リーダー数", "leaders"),
        (69, "ew1 (西大寺勤務）", "ew1"),
        (71, "ew2 (薬師寺勤務)", "ew2"),
        (73, "ew3(吉備勤務)", "ew3"),
        (75, "IW1(クリクラ）", "iw1"),
        (77, "IW2(未設定)", "iw2"),
    ]
    for start_col, label, _ in coverage_sections:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 1)
        ws.cell(2, start_col, label)
        ws.cell(3, start_col, "日勤")
        ws.cell(3, start_col + 1, "夜勤")
        for row in (2, 3):
            for c in (start_col, start_col + 1):
                ws.cell(row, c).fill = header_fill if row == 2 else sub_fill
                ws.cell(row, c).border = border
                ws.cell(row, c).font = Font(bold=True)
                ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    min_day_validation = DataValidation(type="list", formula1='"0,1,2,3,4,5,6,7,8,9,10"', allow_blank=False)
    min_night_validation = DataValidation(type="list", formula1='"0,1,2,3,4,5"', allow_blank=False)
    leader_day_validation = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=False)
    leader_night_validation = DataValidation(type="list", formula1='"0,1,2"', allow_blank=False)
    zero_one_day_validation = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    zero_one_night_validation = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    for dv in (
        min_day_validation,
        min_night_validation,
        leader_day_validation,
        leader_night_validation,
        zero_one_day_validation,
        zero_one_night_validation,
    ):
        ws.add_data_validation(dv)

    for d in range(N_DAYS):
        request_row = 4 + d * 2
        defaults = [
            {"minimum": 3, "leaders": 1, "ew1": 0, "ew2": 0, "ew3": 0, "iw1": 0, "iw2": 0},
            {"minimum": 2, "leaders": 1, "ew1": 0, "ew2": 0, "ew3": 0, "iw1": 0, "iw2": 0},
        ]
        day_coverage = payload.coverage[d] if d < len(payload.coverage) else defaults
        for sh in range(N_SHIFTS):
            values = day_coverage[sh] if sh < len(day_coverage) else defaults[sh]
            for start_col, _, key in coverage_sections:
                cell = ws.cell(request_row, start_col + sh)
                cell.value = int(values.get(key, defaults[sh][key]))
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = holiday_fill if (
                    d < len(payload.holiday_labels) and payload.holiday_labels[d] == "休日"
                ) else PatternFill(fill_type=None)

        min_day_validation.add(ws.cell(request_row, 65))
        min_night_validation.add(ws.cell(request_row, 66))
        leader_day_validation.add(ws.cell(request_row, 67))
        leader_night_validation.add(ws.cell(request_row, 68))
        for start_col in (69, 71, 73, 75, 77):
            zero_one_day_validation.add(ws.cell(request_row, start_col))
            zero_one_night_validation.add(ws.cell(request_row, start_col + 1))

        # 数値ドロップダウンで1以上なら薄いピンク。
        for col in range(65, 79):
            ws.conditional_formatting.add(
                ws.cell(request_row, col).coordinate,
                CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=pink_fill),
            )

        remark_row = request_row + 1
        for col in range(65, 79):
            ws.cell(remark_row, col).border = border
            ws.cell(remark_row, col).fill = remark_fill

    # --------------------------------------------------
    # 管理者設定入力欄（添付Excelの下段レイアウトに合わせる）
    # 66:救急医師カウント / 67:リーダーlevel / 68:適切なシフト数
    # 69-73: EW1 / EW2 / EW3 / IW1 / IW2
    # --------------------------------------------------
    admin_rows = [
        (66, "救急医師カウント"),
        (67, "リーダーlevel"),
        (68, "適切なシフト数"),
        (69, "ew1 (西大寺勤務）"),
        (70, "ew2 (薬師寺勤務)"),
        (71, "ew3 (吉備勤務）"),
        (72, "iw1 (クリクラ）"),
        (73, "iw2 (未設定）"),
    ]
    for row, label in admin_rows:
        ws.cell(row, 1, label)
        for col in range(1, 5):
            ws.cell(row, col).fill = header_fill
            ws.cell(row, col).border = border
            ws.cell(row, col).font = Font(bold=True)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    for i in range(N_STAFF):
        col = 5 + i * 2
        member = payload.staff[i] if i < len(payload.staff) else StaffInput(no=i)
        # 救急医師カウントは管理画面の整数0/1をそのまま出力する。
        ws.cell(66, col, int(member.emergency_count))
        ws.cell(67, col, int(member.leader_level))
        # 適切な勤務数は日勤・夜勤の2セルを使用。
        ws.cell(68, col, int(member.target_day))
        ws.cell(68, col + 1, int(member.target_night))
        # EW1/EW2/EW3/IW2は0/1、IW1は0〜3。
        ws.cell(69, col, 1 if member.ew1_available else 0)
        ws.cell(70, col, 1 if member.ew2_available else 0)
        ws.cell(71, col, 1 if member.ew3_available else 0)
        ws.cell(72, col, int(member.iw1_available))
        ws.cell(73, col, 1 if member.iw2_available else 0)

        for row in range(66, 74):
            for c in (col, col + 1):
                ws.cell(row, c).border = border
                ws.cell(row, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row, c).fill = sub_fill

    # Excel上で誤入力を防ぐため、可否・レベル・適切なシフト数に入力規則を設定。
    binary_validation = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    binary_validation.error = "0 または 1 を入力してください。"
    ws.add_data_validation(binary_validation)
    leader_validation = DataValidation(type="list", formula1='"0,1,2"', allow_blank=False)
    ws.add_data_validation(leader_validation)
    iw1_validation = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=False)
    iw1_validation.error = "0〜3から選択してください。"
    ws.add_data_validation(iw1_validation)
    target_day_validation = DataValidation(
        type="list",
        formula1='"0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20"',
        allow_blank=False,
    )
    target_night_validation = DataValidation(
        type="list",
        formula1='"0,1,2,3,4,5,6,7,8,9,10"',
        allow_blank=False,
    )
    ws.add_data_validation(target_day_validation)
    ws.add_data_validation(target_night_validation)
    for i in range(N_STAFF):
        col = 5 + i * 2
        binary_validation.add(ws.cell(66, col))
        leader_validation.add(ws.cell(67, col))
        target_day_validation.add(ws.cell(68, col))
        target_night_validation.add(ws.cell(68, col + 1))
        for row in (69, 70, 71, 73):
            binary_validation.add(ws.cell(row, col))
        iw1_validation.add(ws.cell(72, col))

        # 整数ドロップダウンで1以上なら薄いピンク。
        for row, c in ((66, col), (67, col), (68, col), (68, col + 1), (69, col), (70, col), (71, col), (72, col), (73, col)):
            ws.conditional_formatting.add(
                ws.cell(row, c).coordinate,
                CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=pink_fill),
            )

    # 左端の長い見出しを読みやすくするため、A:D を結合。
    for row in (1, 2, 3, 66, 67, 68, 69, 70, 71, 72, 73):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --------------------------------------------------
    # 勤務者ごとの境界線を太くする
    # 各勤務者は「日勤・夜勤」の2列なので、夜勤列の右端を太線にする
    # --------------------------------------------------
    for i in range(N_STAFF):
        night_col = 5 + i * 2 + 1

        for row in range(1, 74):
            cell = ws.cell(row=row, column=night_col)

            cell.border = Border(
                left=cell.border.left,
                right=staff_separator,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )
            
    ws.freeze_panes = "E4"
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 10
    for col in range(5, 5 + N_STAFF * 2):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for col in range(65, 79):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, 74):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 24

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return wb


def _parse_request_excel(content: bytes):
    try:
        wb = load_workbook(BytesIO(content), data_only=False)
    except Exception as exc:
        raise HTTPException(400, f"Excelファイルを開けません: {exc}") from exc
    ws = wb["勤務希望入力"] if "勤務希望入力" in wb.sheetnames else wb[wb.sheetnames[0]]

    if ws.max_row < 65 or ws.max_column < 6:
        raise HTTPException(400, "勤務希望Excelの行列数が不足しています。Web画面から出力したExcelを使用してください。")

    staff_names = []
    for i in range(N_STAFF):
        col = 5 + i * 2
        staff_names.append(str(ws.cell(2, col).value or "").strip())

    dates = []
    requests = []
    remarks = []
    for i in range(N_STAFF):
        requests.append([["available", "available"] for _ in range(N_DAYS)])
        remarks.append([["", ""] for _ in range(N_DAYS)])

    for d in range(N_DAYS):
        request_row = 4 + d * 2
        remark_row = request_row + 1
        raw = ws.cell(request_row, 2).value
        if isinstance(raw, datetime):
            parsed = raw.date()
        elif isinstance(raw, date):
            parsed = raw
        elif isinstance(raw, str):
            value = raw.strip().replace("/", "-")
            try:
                parsed = date.fromisoformat(value)
            except ValueError as exc:
                raise HTTPException(400, f"B{request_row} の日付が正しくありません: {raw}") from exc
        else:
            raise HTTPException(400, f"B{request_row} に日付がありません。")
        dates.append(parsed.isoformat())

        for i in range(N_STAFF):
            for sh in range(N_SHIFTS):
                col = 5 + i * 2 + sh
                raw_choice = str(ws.cell(request_row, col).value or "○").strip()
                if raw_choice not in LABEL_TO_REQUEST:
                    coord = ws.cell(request_row, col).coordinate
                    raise HTTPException(
                        400,
                        f"{coord} の勤務希望「{raw_choice}」は選択肢にありません。ドロップダウンから選択してください。",
                    )
                requests[i][d][sh] = LABEL_TO_REQUEST[raw_choice]
                remarks[i][d][sh] = str(ws.cell(remark_row, col).value or "")

    def _int_cell(row: int, col: int, default: int = 0) -> int:
        raw = ws.cell(row, col).value
        if raw in (None, ""):
            return default
        if isinstance(raw, bool):
            return 1 if raw else 0
        try:
            return int(raw)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, f"{ws.cell(row, col).coordinate} は整数で入力してください。") from exc

    # BM:BZ の日別管理者設定を読み込む。
    coverage = []
    coverage_col_map = {
        "minimum": 65,
        "leaders": 67,
        "ew1": 69,
        "ew2": 71,
        "ew3": 73,
        "iw1": 75,
        "iw2": 77,
    }
    for d in range(N_DAYS):
        row = 4 + d * 2
        shifts = []
        for sh in range(N_SHIFTS):
            item = {key: _int_cell(row, col + sh, 0) for key, col in coverage_col_map.items()}
            if sh == 0:
                if not 0 <= item["minimum"] <= 10:
                    raise HTTPException(400, f"{ws.cell(row, 65).coordinate} の最小勤務数（日勤）は0〜10で選択してください。")
                if not 0 <= item["leaders"] <= 3:
                    raise HTTPException(400, f"{ws.cell(row, 67).coordinate} の最小リーダー数（日勤）は0〜3で選択してください。")
            else:
                if not 0 <= item["minimum"] <= 5:
                    raise HTTPException(400, f"{ws.cell(row, 66).coordinate} の最小勤務数（夜勤）は0〜5で選択してください。")
                if not 0 <= item["leaders"] <= 2:
                    raise HTTPException(400, f"{ws.cell(row, 68).coordinate} の最小リーダー数（夜勤）は0〜2で選択してください。")
            for key in ("ew1", "ew2", "ew3", "iw1", "iw2"):
                if item[key] not in (0, 1):
                    coord = ws.cell(row, coverage_col_map[key] + sh).coordinate
                    raise HTTPException(400, f"{coord} の {key.upper()} は0または1で選択してください。")
            shifts.append(item)
        coverage.append(shifts)

    staff = []
    for i in range(N_STAFF):
        col = 5 + i * 2
        emergency_count = _int_cell(66, col, 0)
        leader_level = _int_cell(67, col, 0)
        target_day = _int_cell(68, col, 0)
        target_night = _int_cell(68, col + 1, 0)
        ew1_available = _int_cell(69, col, 0)
        ew2_available = _int_cell(70, col, 0)
        ew3_available = _int_cell(71, col, 0)
        iw1_available = _int_cell(72, col, 0)
        iw2_available = _int_cell(73, col, 0)

        if emergency_count not in (0, 1):
            raise HTTPException(400, f"{ws.cell(66, col).coordinate} の救急医師カウントは0または1で入力してください。")
        if leader_level not in (0, 1, 2):
            raise HTTPException(400, f"{ws.cell(67, col).coordinate} のリーダーlevelは0〜2で入力してください。")
        if not 0 <= target_day <= 20:
            raise HTTPException(400, f"{ws.cell(68, col).coordinate} の適切なシフト数（日勤）は0〜20で選択してください。")
        if not 0 <= target_night <= 10:
            raise HTTPException(400, f"{ws.cell(68, col + 1).coordinate} の適切なシフト数（夜勤）は0〜10で選択してください。")
        for row, value in ((69, ew1_available), (70, ew2_available), (71, ew3_available), (72, iw1_available), (73, iw2_available)):
            if value not in (0, 1):
                raise HTTPException(400, f"{ws.cell(row, col).coordinate} は0または1で入力してください。")

        staff.append({
            "no": i,
            "name": staff_names[i],
            "emergency_count": emergency_count,
            "leader_level": leader_level,
            "ew1_candidate": bool(ew1_available),
            "ew1_available": int(ew1_available),
            "ew2_available": int(ew2_available),
            "ew3_available": int(ew3_available),
            "iw1_available": int(iw1_available),
            "iw2_available": int(iw2_available),
            "target_day": target_day,
            "target_night": target_night,
        })

    return {
        "dates": dates,
        "staff_names": staff_names,
        "staff": staff,
        "requests": requests,
        "remarks": remarks,
        "coverage": coverage,
    }


@app.post("/requests/export")
def export_requests_excel(payload: RequestExcelInput = Body(...)):
    workbook = _request_excel_workbook(payload)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="shift_request_input.xlsx"'},
    )


@app.post("/requests/import")
async def import_requests_excel(file: UploadFile = File(...)):
    if Path(file.filename or "").suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(400, ".xlsx または .xlsm ファイルを選択してください。")
    try:
        content = await file.read()
        if not content:
            raise HTTPException(400, "ファイルが空です。")
        return _parse_request_excel(content)
    finally:
        await file.close()


@app.post("/schedule/solve")
def solve_from_web(payload: ScheduleInput = Body(...)):
    a,work,ew1,iw1=solve_payload(payload); unavailable=a[0]; rows=[]
    for d,raw in enumerate(payload.dates):
        cells=[]
        for s,_ in enumerate(payload.staff):
            values=[]
            for sh in range(N_SHIFTS):
                if unavailable[s,d,sh]: values.append("×")
                elif ew1[s,d,sh]: values.append("外1")
                elif iw1[s,d,sh]: values.append("内1")
                elif work[s,d,sh]: values.append("日" if sh==0 else "夜")
                else: values.append("")
            cells.append(values)
        rows.append({"date":raw,"cells":cells})
    return {"status":"feasible","rows":rows}

@app.post("/schedule/export")
def export_from_web(payload: ScheduleInput = Body(...)):
    a,work,ew1,iw1=solve_payload(payload); unavailable=a[0]; staffno=a[10]; staffname=a[11]; day_week=a[15]
    output=BytesIO(); export_final_work_3d(work,ew1,iw1,unavailable,staffno,staffname,day_week,output); output.seek(0)
    return StreamingResponse(output,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":'attachment; filename="read_work.xlsx"'})

@app.post("/schedule/generate")
async def generate_schedule(file: UploadFile = File(...)):
    if Path(file.filename or "").suffix.lower() not in {".xlsx",".xlsm"}: raise HTTPException(400,".xlsx または .xlsm ファイルを選択してください。")
    try:
        content=await file.read()
        if not content: raise HTTPException(400,"ファイルが空です。")
        v=load_shift_schedule_excel(BytesIO(content)); unavailable,avoid,mandatory,ew1m,iw1m,optimal,min_work,min_leader,ew1c,iw1c,staffno,staffname,leader,is_ew1,iw1p,day_week,pair_ng,_=v
        work,ew1,iw1,status=solve_shift_schedule(unavailable,avoid,mandatory,ew1m,iw1m,optimal,min_work,min_leader,ew1c,iw1c,staffno,leader,is_ew1,iw1p,pair_ng)
        if status not in (cp_model.OPTIMAL,cp_model.FEASIBLE): raise HTTPException(422,"条件を満たすシフトが見つかりません。")
        output=BytesIO(); export_final_work_3d(work,ew1,iw1,unavailable,staffno,staffname,day_week,output); output.seek(0)
        return StreamingResponse(output,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":'attachment; filename="shift_schedule_result.xlsx"'})
    except HTTPException: raise
    except (BadZipFile,InvalidFileException,KeyError,TypeError,ValueError) as exc: raise HTTPException(400,f"Excelファイルの形式または内容が正しくありません: {exc}") from exc
    finally: await file.close()
